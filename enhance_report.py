#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Oct 10 17:14:14 2025

@author: ha
"""
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Oct 10 12:27:28 2025

@author: ha
"""
"""
Enhanced Report Generator - Creates formatted Excel and HTML reports
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# Configuration
PROJECT_FOLDER = '/Users/ha/Documents/gb_report_info/'
INPUT_FILE = 'analyzed_sql_reports.xlsx'
OUTPUT_EXCEL = 'report_info_formatted.xlsx'
OUTPUT_HTML = 'report_info_interactive.html'

def create_formatted_excel():
    """Create a beautifully formatted Excel file"""
    
    os.chdir(PROJECT_FOLDER)
    
    # Read the data
    df = pd.read_excel(INPUT_FILE)
    
    # Save to new file first
    df.to_excel(OUTPUT_EXCEL, index=False, sheet_name='SQL Reports')
    
    # Load workbook for formatting
    wb = load_workbook(OUTPUT_EXCEL)
    ws = wb['SQL Reports']
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    
    alt_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # Format headers
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    
    # Set column widths
    column_widths = {
        'A': 15,  # Main Folder
        'B': 20,  # Folder Path
        'C': 35,  # Report Name
        'D': 15,  # Universe Name
        'E': 20,  # Connection Name
        'F': 15,  # SQL Query (will be hidden)
        'G': 40,  # tables
        'H': 60,  # tbl_fields
        'I': 80   # Description
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Format data rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        # Alternate row colors
        if row_idx % 2 == 0:
            for cell in row:
                cell.fill = alt_row_fill
        
        # Apply borders and alignment
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
        
        # Bold report names
        row[2].font = Font(bold=True, size=10)
    
    # Hide SQL Query column (it's too long)
    ws.column_dimensions['F'].hidden = True
    
    # Freeze the header row
    ws.freeze_panes = 'A2'
    
    # Add auto-filter
    ws.auto_filter.ref = ws.dimensions
    
    # Adjust row heights
    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 60
    
    # Header row height
    ws.row_dimensions[1].height = 30
    
    # Save
    wb.save(OUTPUT_EXCEL)
    print(f"✓ Formatted Excel saved: {OUTPUT_EXCEL}")

def create_html_report():
    """Create an interactive HTML report"""
    
    os.chdir(PROJECT_FOLDER)
    df = pd.read_excel(INPUT_FILE)
    
    # Convert dataframe to JSON properly
    import json
    reports_json = json.dumps(df.to_dict(orient='records'))
    
    html_content = """
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>SQL Reports Analysis</title>
    <style>
        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }
        
        body {
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            padding: 20px;
            min-height: 100vh;
        }
        
        .container {
            max-width: 1400px;
            margin: 0 auto;
            background: white;
            border-radius: 12px;
            box-shadow: 0 20px 60px rgba(0,0,0,0.3);
            overflow: hidden;
        }
        
        .header {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 30px 40px;
        }
        
        .header h1 {
            font-size: 32px;
            margin-bottom: 10px;
        }
        
        .header p {
            font-size: 16px;
            opacity: 0.9;
        }
        
        .stats {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 20px;
            padding: 30px 40px;
            background: #f8f9fa;
            border-bottom: 1px solid #e0e0e0;
        }
        
        .stat-card {
            background: white;
            padding: 20px;
            border-radius: 8px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        }
        
        .stat-card h3 {
            font-size: 14px;
            color: #666;
            text-transform: uppercase;
            letter-spacing: 0.5px;
            margin-bottom: 8px;
        }
        
        .stat-card p {
            font-size: 28px;
            font-weight: bold;
            color: #667eea;
        }
        
        .search-box {
            padding: 20px 40px;
            background: white;
            border-bottom: 1px solid #e0e0e0;
        }
        
        .search-box input {
            width: 100%;
            padding: 12px 20px;
            font-size: 16px;
            border: 2px solid #e0e0e0;
            border-radius: 8px;
            outline: none;
            transition: all 0.3s;
        }
        
        .search-box input:focus {
            border-color: #667eea;
            box-shadow: 0 0 0 3px rgba(102, 126, 234, 0.1);
        }
        
        .filters {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 15px;
            padding: 20px 40px;
            background: white;
            border-bottom: 1px solid #e0e0e0;
        }
        
        .filter-group {
            display: flex;
            flex-direction: column;
        }
        
        .filter-label {
            font-size: 12px;
            font-weight: 600;
            color: #666;
            margin-bottom: 6px;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }
        
        .filter-select {
            padding: 10px;
            border: 2px solid #e0e0e0;
            border-radius: 6px;
            font-size: 14px;
            outline: none;
            transition: all 0.3s;
            background: white;
        }
        
        .filter-select:focus {
            border-color: #667eea;
            box-shadow: 0 0 0 3px rgba(102, 126, 234, 0.1);
        }
        
        .reports {
            padding: 40px;
        }
        
        .report-card {
            background: white;
            border: 1px solid #e0e0e0;
            border-radius: 8px;
            margin-bottom: 20px;
            overflow: hidden;
            transition: all 0.3s;
        }
        
        .report-card:hover {
            box-shadow: 0 4px 12px rgba(0,0,0,0.1);
            transform: translateY(-2px);
        }
        
        .report-header {
            padding: 20px 25px;
            background: linear-gradient(to right, #f8f9fa, white);
            border-bottom: 1px solid #e0e0e0;
            cursor: pointer;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }
        
        .report-header:hover {
            background: linear-gradient(to right, #e9ecef, #f8f9fa);
        }
        
        .report-title {
            font-size: 20px;
            font-weight: 600;
            color: #333;
            display: flex;
            align-items: center;
            gap: 10px;
        }
        
        .report-number {
            display: inline-block;
            background: #667eea;
            color: white;
            width: 32px;
            height: 32px;
            border-radius: 50%;
            text-align: center;
            line-height: 32px;
            font-size: 14px;
            font-weight: bold;
        }
        
        .expand-icon {
            font-size: 24px;
            color: #667eea;
            transition: transform 0.3s;
        }
        
        .report-card.expanded .expand-icon {
            transform: rotate(180deg);
        }
        
        .report-content {
            max-height: 0;
            overflow: hidden;
            transition: max-height 0.3s ease-out;
        }
        
        .report-card.expanded .report-content {
            max-height: 2000px;
        }
        
        .report-details {
            padding: 25px;
        }
        
        .detail-section {
            margin-bottom: 25px;
        }
        
        .detail-section:last-child {
            margin-bottom: 0;
        }
        
        .detail-label {
            font-size: 12px;
            font-weight: 600;
            text-transform: uppercase;
            color: #667eea;
            letter-spacing: 0.5px;
            margin-bottom: 8px;
        }
        
        .detail-value {
            font-size: 15px;
            color: #333;
            line-height: 1.6;
            background: #f8f9fa;
            padding: 12px 15px;
            border-radius: 6px;
            border-left: 3px solid #667eea;
        }
        
        .tags {
            display: flex;
            flex-wrap: wrap;
            gap: 8px;
            margin-top: 8px;
        }
        
        .tag {
            display: inline-block;
            background: #e3f2fd;
            color: #1976d2;
            padding: 6px 12px;
            border-radius: 4px;
            font-size: 13px;
            font-weight: 500;
        }
        
        .sql-query {
            background: #1e1e1e;
            color: #d4d4d4;
            padding: 20px;
            border-radius: 6px;
            overflow-x: auto;
            font-family: 'Consolas', 'Monaco', monospace;
            font-size: 13px;
            line-height: 1.6;
            white-space: pre-wrap;
        }
        
        .sql-container {
            position: relative;
        }
        
        .copy-btn {
            position: absolute;
            top: 10px;
            right: 10px;
            background: #667eea;
            color: white;
            border: none;
            padding: 8px 16px;
            border-radius: 6px;
            cursor: pointer;
            font-size: 12px;
            font-weight: 600;
            transition: all 0.3s;
            z-index: 10;
        }
        
        .copy-btn:hover {
            background: #5568d3;
            transform: translateY(-2px);
            box-shadow: 0 4px 8px rgba(102, 126, 234, 0.3);
        }
        
        .copy-btn.copied {
            background: #10b981;
        }
        
        .no-results {
            text-align: center;
            padding: 60px 20px;
            color: #999;
        }
        
        .no-results h2 {
            font-size: 24px;
            margin-bottom: 10px;
        }
        
        @media (max-width: 768px) {
            body {
                padding: 10px;
            }
            
            .header {
                padding: 20px;
            }
            
            .header h1 {
                font-size: 24px;
            }
            
            .reports {
                padding: 20px;
            }
            
            .report-title {
                font-size: 16px;
            }
        }
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>📊 SQL Reports Analysis</h1>
            <p>AML/FCUBS Database Reports - Comprehensive Analysis</p>
        </div>
        
        <div class="stats">
            <div class="stat-card">
                <h3>Total Reports</h3>
                <p id="total-reports">0</p>
            </div>
            <div class="stat-card">
                <h3>Total Tables</h3>
                <p id="total-tables">0</p>
            </div>
            <div class="stat-card">
                <h3>Main Category</h3>
                <p id="main-category" style="font-size: 20px;">-</p>
                <p style="font-size: 12px; color: #999; margin-top: 5px;" id="category-count">-</p>
            </div>
        </div>
    
        
        <div class="search-box">
            <input type="text" id="searchInput" placeholder="🔍 Search reports by name, table, or description..." />
        </div>
        
        <div class="filters">
            <div class="filter-group">
                <label class="filter-label">Folder</label>
                <select id="folderFilter" class="filter-select">
                    <option value="">All Folders</option>
                </select>
            </div>
            <div class="filter-group">
                <label class="filter-label">Connection</label>
                <select id="connectionFilter" class="filter-select">
                    <option value="">All Connections</option>
                </select>
            </div>
            <div class="filter-group">
                <label class="filter-label">Table Count</label>
                <select id="tableCountFilter" class="filter-select">
                    <option value="">All</option>
                    <option value="1-3">1-3 tables</option>
                    <option value="4-6">4-6 tables</option>
                    <option value="7+">7+ tables</option>
                </select>
            </div>
        </div>
        
        <div class="reports" id="reportsContainer">
            <!-- Reports will be inserted here -->
        </div>
    </div>
    
    <script>
        const reportsData = """ + reports_json + """;
        
        function renderReports(reports) {
            const container = document.getElementById('reportsContainer');
            container.innerHTML = '';
            
            if (reports.length === 0) {
                container.innerHTML = `
                    <div class="no-results">
                        <h2>No reports found</h2>
                        <p>Try adjusting your search terms or filters</p>
                    </div>
                `;
                return;
            }
            
            reports.forEach((report, index) => {
                const tables = report.tables ? report.tables.split(',').map(t => t.trim()) : [];
                
                const reportCard = document.createElement('div');
                reportCard.className = 'report-card';
                reportCard.innerHTML = `
                    <div class="report-header" onclick="toggleReport(this)">
                        <div class="report-title">
                            <span class="report-number">${index + 1}</span>
                            ${escapeHtml(report['Report Name'] || 'Unnamed Report')}
                        </div>
                        <span class="expand-icon">▼</span>
                    </div>
                    <div class="report-content">
                        <div class="report-details">
                            <div class="detail-section">
                                <div class="detail-label">Description</div>
                                <div class="detail-value">${escapeHtml(report.Description || 'No description available')}</div>
                            </div>
                            
                            <div class="detail-section">
                                <div class="detail-label">Tables Used (${tables.length})</div>
                                <div class="tags">
                                    ${tables.map(table => `<span class="tag">${escapeHtml(table)}</span>`).join('')}
                                </div>
                            </div>
                            
                            <div class="detail-section">
                                <div class="detail-label">Table Fields</div>
                                <div class="detail-value">${escapeHtml(report.tbl_fields || 'No field information available')}</div>
                            </div>
                            
                            <div class="detail-section">
                                <div class="detail-label">Connection</div>
                                <div class="detail-value">${escapeHtml(report['Connection Name'] || 'N/A')}</div>
                            </div>
                            
                            <div class="detail-section">
                                <div class="detail-label">Folder Path</div>
                                <div class="detail-value">${escapeHtml(report['Folder Path'] || 'N/A')}</div>
                            </div>
                            
                            <div class="detail-section">
                                <div class="detail-label">SQL Query</div>
                                <div class="sql-container">
                                    <button class="copy-btn" onclick="copySQL(this, ${index})">📋 Copy SQL</button>
                                    <div class="sql-query" id="sql-${index}">${escapeHtml(report['SQL Query'] || 'No query available')}</div>
                                </div>
                            </div>
                        </div>
                    </div>
                `;
                container.appendChild(reportCard);
            });
        }
        
        function toggleReport(element) {
            const card = element.closest('.report-card');
            card.classList.toggle('expanded');
        }
        
        function escapeHtml(text) {
            if (!text) return '';
            const div = document.createElement('div');
            div.textContent = text;
            return div.innerHTML;
        }
        
        function copySQL(buttonElement, index) {
            const sqlText = reportsData[index]['SQL Query'];
            
            navigator.clipboard.writeText(sqlText).then(() => {
                const originalText = buttonElement.innerHTML;
                buttonElement.innerHTML = '✓ Copied!';
                buttonElement.classList.add('copied');
                
                setTimeout(() => {
                    buttonElement.innerHTML = originalText;
                    buttonElement.classList.remove('copied');
                }, 2000);
            }).catch(err => {
                alert('Failed to copy: ' + err);
            });
        }
                

        
        function populateFilters() {
            const folders = new Set();
            const connections = new Set();
            
            reportsData.forEach(report => {
                if (report['Folder Path']) folders.add(report['Folder Path']);
                if (report['Connection Name']) connections.add(report['Connection Name']);
            });
            
            const folderFilter = document.getElementById('folderFilter');
            Array.from(folders).sort().forEach(folder => {
                const option = document.createElement('option');
                option.value = folder;
                option.textContent = folder;
                folderFilter.appendChild(option);
            });
            
            const connectionFilter = document.getElementById('connectionFilter');
            Array.from(connections).sort().forEach(conn => {
                const option = document.createElement('option');
                option.value = conn;
                option.textContent = conn;
                connectionFilter.appendChild(option);
            });
        }
        
        function applyFilters() {
            const searchTerm = document.getElementById('searchInput').value.toLowerCase();
            const folderFilter = document.getElementById('folderFilter').value;
            const connectionFilter = document.getElementById('connectionFilter').value;
            const tableCountFilter = document.getElementById('tableCountFilter').value;
            
            const filtered = reportsData.filter(report => {
                const matchesSearch = !searchTerm || (
                    (report['Report Name'] && report['Report Name'].toLowerCase().includes(searchTerm)) ||
                    (report.Description && report.Description.toLowerCase().includes(searchTerm)) ||
                    (report.tables && report.tables.toLowerCase().includes(searchTerm))
                );
                
                const matchesFolder = !folderFilter || report['Folder Path'] === folderFilter;
                const matchesConnection = !connectionFilter || report['Connection Name'] === connectionFilter;
                
                let matchesTableCount = true;
                if (tableCountFilter && report.tables) {
                    const tableCount = report.tables.split(',').length;
                    if (tableCountFilter === '1-3') {
                        matchesTableCount = tableCount >= 1 && tableCount <= 3;
                    } else if (tableCountFilter === '4-6') {
                        matchesTableCount = tableCount >= 4 && tableCount <= 6;
                    } else if (tableCountFilter === '7+') {
                        matchesTableCount = tableCount >= 7;
                    }
                }
                
                return matchesSearch && matchesFolder && matchesConnection && matchesTableCount;
            });
            
            renderReports(filtered);
        }
        
        // Event listeners
        document.getElementById('searchInput').addEventListener('input', applyFilters);
        document.getElementById('folderFilter').addEventListener('change', applyFilters);
        document.getElementById('connectionFilter').addEventListener('change', applyFilters);
        document.getElementById('tableCountFilter').addEventListener('change', applyFilters);
        
        // Calculate and display stats
        document.getElementById('total-reports').textContent = reportsData.length;
        const allTables = new Set();
        reportsData.forEach(report => {
            if (report.tables) {
                report.tables.split(',').forEach(table => allTables.add(table.trim()));
            }
        });
        document.getElementById('total-tables').textContent = allTables.size;
        
        // Calculate main category dynamically
        const folderCounts = {};
        reportsData.forEach(report => {
            const folder = report['Folder Path'];
            if (folder) {
                folderCounts[folder] = (folderCounts[folder] || 0) + 1;
            }
        });
        
        if (Object.keys(folderCounts).length > 0) {
            const mainCategory = Object.keys(folderCounts).reduce((a, b) => 
                folderCounts[a] > folderCounts[b] ? a : b
            );
            const categoryName = mainCategory.split('/').pop(); // Get last part after /
            document.getElementById('main-category').textContent = categoryName;
            document.getElementById('category-count').textContent = 
                Object.keys(folderCounts).length + ' total categories';
        } else {
            document.getElementById('main-category').textContent = 'N/A';
            document.getElementById('category-count').textContent = 'No data';
        }
            
        
        // Initialize
        populateFilters();
        renderReports(reportsData);
    </script>
</body>
</html>
    """
    
    with open(OUTPUT_HTML, 'w', encoding='utf-8') as f:
        f.write(html_content)
    
    print(f"✓ Interactive HTML saved: {OUTPUT_HTML}")
    print(f"✓ Open in browser: file://{os.path.join(PROJECT_FOLDER, OUTPUT_HTML)}")

def main():
    """Generate both reports"""
    print("=" * 70)
    print("Creating Enhanced Reports")
    print("=" * 70)
    
    print("\n1. Creating formatted Excel...")
    create_formatted_excel()
    
    print("\n2. Creating interactive HTML...")
    create_html_report()
    
    print("\n" + "=" * 70)
    print("✅ All reports created successfully!")
    print("=" * 70)
    print(f"\n📄 Formatted Excel: {OUTPUT_EXCEL}")
    print(f"🌐 Interactive HTML: {OUTPUT_HTML}")
    print("\nOpen the HTML file in your browser for the best experience!")

if __name__ == "__main__":
    main()